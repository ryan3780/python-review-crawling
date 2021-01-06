import time
import datetime
import os

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook, load_workbook



options = Options()
options.headless = True
browser = webdriver.Chrome(executable_path="./chromedriver.exe", options=options)
url = 'https://galaxystore.samsung.com/detail/com.monotype.android.font.onlydoremik'

now = datetime.datetime.now()
date = now.strftime('%Y.%m.%d')


excel_file_path = os.getcwd() + '/log/'
excel_file_name = excel_file_path + url.split('font.')[1] + '_' + date + '.xlsx'
excel_sheet_title = 'reviews'


def make_excel():
    work_book = Workbook()
    sheet1 = work_book.active
    sheet1.title = excel_sheet_title

    sheet1.cell(row= 2, column= 2).value = '번호'
    sheet1.cell(row= 2, column= 3).value = '날짜'
    sheet1.cell(row= 2, column= 4).value = '별점'
    sheet1.cell(row= 2, column= 5).value = '내용'

    work_book.save(filename = excel_file_name)
    work_book.close()


make_excel()

browser.get(url)

time.sleep(3)

tag_names = browser.find_element_by_css_selector(".Tabs_tab_list__ge63w").find_elements_by_tag_name("li")

def clickReviews():
    reviews = browser.find_element_by_css_selector(".Tabs_tab_list__ge63w").find_elements_by_tag_name("li")[1]
    reviews.click()

clickReviews()

time.sleep(3)

star_list = []
date_list = []
review_list = []
crawling_results = []

def clickReadMoreReviews():

    try:
        read_more_reviews = browser.find_element_by_css_selector(".CustomReviewContainer_morereview_btn__3xmKM")
        read_more_reviews.click()
        time.sleep(3)
        clickReadMoreReviews()
    except:
        time.sleep(3)
        reviews = browser.find_elements_by_css_selector('.CustomerReview_customer_review__qmUeI')
        stars = browser.find_elements_by_css_selector('.RatingStar_star__1Va97')
        

        for jdx, val in enumerate(stars):
            star_cnt = val.size['width']
            if star_cnt <= 8:
                star_list.append('0.5개')
            elif star_cnt > 8 and star_cnt <= 16:
                star_list.append('1개')
            elif star_cnt > 16 and star_cnt <= 24:
                star_list.append('1.5개')    
            elif star_cnt > 24 and star_cnt <= 32:
                star_list.append('2개')
            elif star_cnt > 32 and star_cnt <= 40:
                star_list.append('2.5개')
            elif star_cnt > 40 and star_cnt <= 48:
                star_list.append('3개')
            elif star_cnt > 48 and star_cnt <= 56:             
                star_list.append('3.5개')   
            elif star_cnt > 56 and star_cnt <= 64:         
                star_list.append('4개')
            elif star_cnt > 64 and star_cnt <= 72:         
                star_list.append('4.5개')
            elif star_cnt > 72 and star_cnt <= 81:
                star_list.append('5개')
    

        result = open('all_reviews.txt', 'w', encoding='utf-8')

        for idx, val in enumerate(reviews):
            
            # print(idx + 1, star_list[idx + 1] , val.text.split("\n"))
            date_list.append(str(val.text.split("\n")[1]))
            review_list.append(str(val.text.split("\n")[2]))
            result.write(str(idx + 1) +'번째 댓글 : '+ star_list[idx + 1] + ' ' + date_list[idx + 1] + ' ' +review_list[idx + 1] + '\n')
            print(str(idx + 1) +'번째 댓글 : '+ star_list[idx + 1] + ' ' +  date_list[idx + 1] + ' ' + review_list[idx + 1] + '\n')
            crawling_results.append([idx + 1, date_list[idx + 1], star_list[idx + 1], review_list[idx + 1] ])

        result.close()    
        

clickReadMoreReviews()


def insert_data_to_excel(crawling_results):
    excel_file = load_workbook(excel_file_name)
    sheet1 = excel_file[excel_sheet_title]

    excel_row = 3
    for idx,data in enumerate(crawling_results):
        sheet1.cell(row=excel_row , column= 2 ).value = data[0]
        sheet1.cell(row=excel_row , column= 3 ).value = data[1]
        sheet1.cell(row=excel_row , column= 4 ).value = data[2]
        sheet1.cell(row=excel_row , column= 5 ).value = data[3]
        excel_row += 1

    excel_file.save(excel_file_name)
    excel_file.close()


insert_data_to_excel(crawling_results)


# print(len(crawling_results))
# print(len(star_list))